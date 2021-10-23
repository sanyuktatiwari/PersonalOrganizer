from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys
from PyQt5.Qt import QTabWidget, QWidget, QFormLayout, QLineEdit, QApplication,\
    QPushButton, QTextEdit, QComboBox
import functools
import operator
from functools import partial
from openpyxl import Workbook
import glob
import openpyxl
from datetime import date
from datetime import datetime
import string
from torch._C import NoneType
import os
from PyQt5 import QtCore
from PyQt5.QtCore import *
    
class tabFrame(QTabWidget):
    def __init__(self, parent = None):
        super(tabFrame,self).__init__(parent)
        self.createTab = QWidget()
        self.insertTab = QWidget()
        #self.editTab = QWidget()
        self.deleteTab = QWidget()
        self.viewTab = QWidget()
        
        self.addTab(self.createTab, "CREATE")
        self.addTab(self.insertTab, "INSERT")
        #self.addTab(self.editTab, "EDIT")
        self.addTab(self.deleteTab,"DELETE")
        self.addTab(self.viewTab, "VIEW")
        
        self.createTabUI()
        self.insertTabUI()
        #self.editTabUI()
        self.deleteTabUI()
        self.viewTabUI()
        
        self.setWindowTitle("PERSONAL ORGANIZER")
        self.resize(500,300)

        #self.currentChanged.connect(self, QtCore. SIGNAL(currentChanged(int)), this, SLOT(self.tabChangedSlot()))
        
    def createTabUI(self):
        layout = QFormLayout()
        #Text Edit to enter the name of new Worksheet to create
        self.newWorkbookName = QTextEdit()
        layout.addRow("Name your Workbook", self.newWorkbookName)
        #Push button
        pushButton = QPushButton()
        pushButton.setText("ADD Workbook")
        layout.addRow("", pushButton)
        pushButton.clicked.connect(self.onclick)
        #Text Edit to show list of available Worksheets
        self.listWorkbooks = QTextEdit()
        self.listWorkbooks.setReadOnly(True)
        self.listWorkbooks.setPlainText(self.availableWorkbooks())
        layout.addRow("List of available Work sheets", self.listWorkbooks)
        
        self.createTab.setLayout(layout)
        
    def onclick(self):
        wb = Workbook()
        wb.save(filename = self.newWorkbookName.toPlainText()+".xlsx")
        self.listWorkbooks.setPlainText(self.availableWorkbooks())
        self.newWorkbookName.clear()
        
    def availableWorkbooks(self):
        workbooks = []
        for file in glob.glob("*.xlsx"):
            workbooks.append(file)
        print(''.join(str(workbook) for workbook in workbooks))
        return ' \n'.join(str(workbook) for workbook in workbooks)
    
    def insertTabUI(self):
        layout = QFormLayout()
        self.dropDownWBList = QComboBox()
        workbooks = []
        for file in glob.glob("*.xlsx"):
            workbooks.append(file)
        for workbook in workbooks:
            self.dropDownWBList.addItem(workbook)
        layout.addRow("Workbooks available", self.dropDownWBList)
        
        #Insert Text Edit to enter the row
        self.textToEnter = QTextEdit()
        layout.addRow("Enter Text",self.textToEnter)
        
        #Push button
        pushButton = QPushButton()
        pushButton.setText("INSERT")
        layout.addRow("", pushButton)
        pushButton.clicked.connect(self.onInsertTextclick)
        layout.addRow("Enter Text to insert", pushButton)
        
        self.insertTab.setLayout(layout)
        
    def onInsertTextclick(self):
        text = self.textToEnter.toPlainText()
        wb = openpyxl.load_workbook(str(self.dropDownWBList.currentText()))
        sheet = wb.active
        sheet.append((date.today(),text))
        print(text)
        self.textToEnter.clear()
        wb.save(str(self.dropDownWBList.currentText()))
        wb.close()
        
    def viewTabUI(self):
        layout = QFormLayout()
        self.dropDownWBList = QComboBox()
        workbooks = []
        for file in glob.glob("*.xlsx"):
            workbooks.append(file)
        for workbook in workbooks:
            self.dropDownWBList.addItem(workbook)
        layout.addRow("Workbooks to view", self.dropDownWBList)
        
        self.textToView = QTextEdit()
        self.viewData()
        self.dropDownWBList.currentIndexChanged.connect(self.onViewComboBoxSelected)
        
        layout.addRow("Worksheet Data",self.textToView)
        
        self.viewTab.setLayout(layout)
        
    def onViewComboBoxSelected(self):
       self.viewData()
        
    def viewData(self):
        #View Text Edit Field
        data = ""
        wb = openpyxl.load_workbook(str(self.dropDownWBList.currentText()))
        sheet = wb.active
        for value in sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=3, values_only=True):
            for entry in value:
                if type(entry) == datetime:
                    data = data + entry.strftime("%m/%d/%Y") + "\t"
                elif type(entry) == str:
                    data = data + entry
            data = data + "\n"
        self.textToView.setPlainText(data)
        
    def deleteTabUI(self):
        layout = QFormLayout()
        self.dropDownWBList = QComboBox()
        workbooks = []
        for file in glob.glob("*.xlsx"):
            workbooks.append(file)
        for workbook in workbooks:
            self.dropDownWBList.addItem(workbook)
        layout.addRow("Workbooks available", self.dropDownWBList)
        
        #Push button
        pushButton = QPushButton()
        pushButton.setText("DELETE")
        layout.addRow("", pushButton)
        pushButton.clicked.connect(self.onDeleteTextclick)
        layout.addRow("Delete Workbook", pushButton)
        
        self.deleteTab.setLayout(layout)
        
    def onDeleteTextclick(self):
        os.remove(str(self.dropDownWBList.currentText()))
        
    def tabChangedSlot(self):
        self.availableWorkbooks()

def main():
    app = QApplication(sys.argv)
    program = tabFrame()
    program.show()
    sys.exit(app.exec_())
    
if __name__ == '__main__':
    main()